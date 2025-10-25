import streamlit as st
import pandas as pd
import re, io, datetime, os, subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
CURRENCY="EUR"; CURRENCY_FORMAT='#,##0.00 "EUR"'; PURCHASE_LOCATION_CODE="KAUNAS"; ALLOC_LOCATION_CODE="KAUNAS"
def get_app_version():
    try:
        cnt=subprocess.check_output(["git","rev-list","--count","HEAD"],stderr=subprocess.DEVNULL).decode().strip(); sha=subprocess.check_output(["git","rev-parse","--short","HEAD"],stderr=subprocess.DEVNULL).decode().strip(); return f"v{int(cnt):03d} ({sha})"
    except Exception:
        env_ver=os.getenv("APP_VERSION") or os.getenv("COMMIT_SHA"); return env_ver if env_ver else "v000"
def coalesce_cols(df,target,candidates):
    if target not in df.columns: df[target]=""
    for c in candidates:
        if c in df.columns: df[target]=df[target].where(df[target].notna()&(df[target].astype(str).str.strip()!=""),df[c])
    return df.drop(columns=[c for c in candidates if c in df.columns],errors="ignore")
def ensure_scalar_strings(df):
    import numpy as np
    def _to_scalar(v):
        if isinstance(v,pd.Series): v=v.dropna(); return "" if v.empty else _to_scalar(v.iloc[0])
        if isinstance(v,(list,tuple,set,np.ndarray,dict)): return str(v)
        return v
    return df.applymap(_to_scalar)
def safe_parse_qty(x):
    if pd.isna(x): return 0.0
    if isinstance(x,(int,float)): return float(x)
    s=str(x).strip()
    if s in {"-","–","—",""}: return 0.0
    s=s.replace("\\xa0","").replace(" ","")
    if "," in s and "." in s: s=s.replace(",","")
    else: s=s.replace(".","").replace(",",".")
    try: return float(s)
    except Exception: return 0.0
def get_excluded_from_stock(df_stock):
    if df_stock is None or df_stock.empty: return set(),set()
    cols=list(df_stock.columns)
    if len(cols)<3: return set(),set()
    s=df_stock.rename(columns={cols[0]:"Component",cols[2]:"Comment"}); m=s["Comment"].astype(str).str.strip().str.lower().isin(["no need","q1"]); comp=s.loc[m,"Component"].astype(str)
    by_type=comp.str.upper().str.replace(" ","").str.strip()
    def _norm_no(x):
        try: return str(int(float(str(x).replace(",","." ).strip())))
        except: return str(x).strip()
    by_no=comp.map(_norm_no); return set(by_type),set(by_no)
def add_extra_components(df,extras):
    if df is None: df=pd.DataFrame()
    out=df.copy()
    for e in extras or []: out=pd.concat([out,pd.DataFrame([{"Original Type":e.get("type",""),"Quantity":e.get("qty",1),"Source":"Extra","No.":e.get("force_no",e.get("type",""))}])],ignore_index=True)
    return out
def build_nav_table_from_bom(df_bom,df_part_no,label="Project BOM"):
    req={"PartNo_A","SupplierNo_E","Manufacturer_D"}
    if df_part_no is None or df_part_no.empty or not req.issubset(df_part_no.columns): return pd.DataFrame(columns=["Entry Type","No.","Quantity","Supplier","Profit","Discount","Description"])
    supplier_map=dict(zip(df_part_no["PartNo_A"].astype(str),df_part_no["SupplierNo_E"])); manuf_map=dict(zip(df_part_no["PartNo_A"].astype(str),df_part_no["Manufacturer_D"].astype(str)))
    tmp=df_bom.copy(); 
    if "Quantity" not in tmp: tmp["Quantity"]=0
    if "Description" not in tmp: tmp["Description"]=""
    if "No." not in tmp: tmp["No."]=""
    tmp["No."]=tmp["No."].astype(str); tmp["Quantity"]=pd.to_numeric(tmp["Quantity"],errors="coerce").fillna(0); rows=[]
    for _,r in tmp.iterrows():
        part_no=str(r["No."]).strip(); qty=safe_parse_qty(r.get("Quantity",0)); manuf=manuf_map.get(part_no,""); profit=10 if "DANFOSS" in str(manuf).upper() else 17; supplier=supplier_map.get(part_no,30093)
        rows.append({"Entry Type":"Item","No.":part_no,"Quantity":qty,"Supplier":supplier,"Profit":profit,"Discount":0,"Description":r.get("Description","")})
    return pd.DataFrame(rows,columns=["Entry Type","No.","Quantity","Supplier","Profit","Discount","Description"])
def pipeline_1_1_norm_name(x): return "".join(str(x).upper().split())
def pipeline_1_2_parse_qty(x): return safe_parse_qty(x)
def pipeline_1_4_normalize_no(x):
    try: return str(int(float(str(x).replace(",","." ).strip())))
    except Exception: return str(x).strip()
def read_excel_any(file,**kwargs):
    try: return pd.read_excel(file,engine="openpyxl",**kwargs)
    except Exception: return pd.read_excel(file,engine="xlrd",**kwargs)
def allocate_from_stock(no,qty_needed,stock_rows):
    allocations=[]; qty_needed=float(pd.to_numeric(pd.Series([qty_needed]),errors="coerce").fillna(0).iloc[0]); remaining=qty_needed
    if stock_rows is not None and not stock_rows.empty:
        for _,srow in stock_rows.iterrows():
            if remaining<=0: break
            bin_code=str(srow.get("Bin Code","")).strip(); stock_qty=float(pd.to_numeric(pd.Series([srow.get("Quantity",0)]),errors="coerce").fillna(0).iloc[0])
            if stock_qty<=0: continue
            if bin_code=="67-01-01-01": continue
            take=min(stock_qty,remaining)
            if take>0: allocations.append({"No.":no,"Bin Code":bin_code,"Allocated Qty":take}); remaining-=take
    if remaining>0: allocations.append({"No.":no,"Bin Code":"","Allocated Qty":remaining})
    return allocations
normalize_no=pipeline_1_4_normalize_no
def pipeline_2_1_user_inputs():
    st.subheader("Project Information")
    pn = st.text_input("Project number (1234-567)", help="Use 4 digits, dash, 3 digits. Regular -, en– or em— dash allowed.")
    norm_pn = re.sub(r"\s*[-–—]\s*", "-", (pn or "").strip())
    if norm_pn and not re.match(r"^\d{4}-\d{3}$", norm_pn):
        st.error("Invalid format (must be 1234-567)")
        return None
    types=["A","B","B1","B2","C","C1","C2","C3","C4","C4.1","C5","C6","C7","C8","F","F1","F2","F3","F4","F4.1","F5","F6","F7","G","G1","G2","G3","G4","G5","G6","G7","Custom"]
    switches=["C160S4FM","C125S4FM","C080S4FM","31115","31113","31111","31109","31107","C404400S","C634630S"]
    return {"project_number":norm_pn,"panel_type":st.selectbox("Panel type",types),"grounding":st.selectbox("Grounding type",["TT","TN-S","TN-C-S"]),"main_switch":st.selectbox("Main switch",switches),"swing_frame":st.checkbox("Swing frame?"),"ups":st.checkbox("UPS?"),"rittal":st.checkbox("Rittal?")}
# REPLACE pipeline_2_2_file_uploads(...) WITH THIS VERSION
def pipeline_2_2_file_uploads(rittal=False):
    st.subheader("Upload Required Files")
    uploads = st.session_state.setdefault("uploads", {})
    dfs = {}
    def _read_cached(key, *, skiprows=None):
        if key in uploads and uploads[key]:
            try: return read_excel_any(io.BytesIO(uploads[key]), skiprows=skiprows) if skiprows is not None else read_excel_any(io.BytesIO(uploads[key]))
            except Exception: return read_excel_any(io.BytesIO(uploads[key]))
        return None
    if not rittal:
        up_cubic = st.file_uploader("Insert CUBIC BOM", type=["xls","xlsx","xlsm"], key="cubic_bom")
        if up_cubic: uploads["cubic_bom"] = up_cubic.getvalue()
        df_cubic = _read_cached("cubic_bom", skiprows=15)
        if df_cubic is not None:
            df_cubic = df_cubic.rename(columns=lambda c: str(c).strip())
            qty_cols = [c for c in df_cubic.columns if str(c).strip() in {"E","F","G"}]
            combo_cols = [c for c in df_cubic.columns if re.sub(r"\s+","",str(c)).upper() in {"E+F+G","E+F","F+G","E+G"} or (("E" in str(c).upper()) and ("F" in str(c).upper()) and ("G" in str(c).upper()))]
            if qty_cols: df_cubic["Quantity"] = df_cubic[qty_cols].bfill(axis=1).iloc[:,0]
            elif combo_cols:
                cc = combo_cols[0]
                df_cubic["Quantity"] = df_cubic[cc].apply(lambda v: safe_parse_qty(re.search(r"([0-9]+[.,]?[0-9]*)", str(v)).group(1)) if (pd.notna(v) and re.search(r"([0-9]+[.,]?[0-9]*)", str(v))) else 0.0)
            else:
                if "Quantity" not in df_cubic.columns: df_cubic["Quantity"] = 0
            df_cubic["Quantity"] = pd.to_numeric(df_cubic["Quantity"], errors="coerce").fillna(0)
            if "Item Id" in df_cubic.columns: df_cubic = df_cubic.rename(columns={"Item Id":"Original Type"})
            else: df_cubic["Original Type"] = df_cubic[df_cubic.columns[0]].astype(str)
            if "No." not in df_cubic.columns: df_cubic["No."] = df_cubic["Original Type"]
            dfs["cubic_bom"] = df_cubic
    up_bom = st.file_uploader("Insert BOM", type=["xls","xlsx","xlsm"], key="bom")
    if up_bom: uploads["bom"] = up_bom.getvalue()
    df_bom = _read_cached("bom")
    if df_bom is not None:
        if df_bom.shape[1] >= 2:
            colA = df_bom.iloc[:,0].fillna("").astype(str).str.strip()
            colB = df_bom.iloc[:,1].fillna("").astype(str).str.strip()
            df_bom["Original Article"] = colA; df_bom["Original Type"] = colB.where(colB!="", colA)
        else:
            df_bom["Original Article"] = df_bom.iloc[:,0].fillna("").astype(str).str.strip()
            df_bom["Original Type"] = df_bom["Original Article"]
        dfs["bom"] = df_bom
    up_data = st.file_uploader("Insert DATA", type=["xls","xlsx","xlsm"], key="data")
    if up_data: uploads["data"] = up_data.getvalue()
    if "data" in uploads and uploads["data"]:
        dfs["data"] = pd.read_excel(io.BytesIO(uploads["data"]), sheet_name=None)
    up_ks = st.file_uploader("Insert Kaunas Stock", type=["xls","xlsx","xlsm"], key="ks")
    if up_ks: uploads["ks"] = up_ks.getvalue()
    df_ks = _read_cached("ks")
    if df_ks is not None: dfs["ks"] = df_ks
    return dfs

def pipeline_2_3_get_sheet_safe(data_dict,names):
    if not isinstance(data_dict,dict): return None
    targets=[n.upper().replace(" ","_") for n in names]
    for key in data_dict.keys():
        if str(key).strip().upper().replace(" ","_") in targets: return data_dict[key]
    return None
def pipeline_2_4_normalize_part_no(df_raw):
    if df_raw is None or df_raw.empty: return pd.DataFrame()
    df=df_raw.copy().rename(columns=lambda c:str(c).strip()); col_map={}
    if df.shape[1]>=1: col_map[df.columns[0]]="PartNo_A"
    if df.shape[1]>=2: col_map[df.columns[1]]="PartName_B"
    if df.shape[1]>=3: col_map[df.columns[2]]="Desc_C"
    if df.shape[1]>=4: col_map[df.columns[3]]="Manufacturer_D"
    if df.shape[1]>=5: col_map[df.columns[4]]="SupplierNo_E"
    if df.shape[1]>=6: col_map[df.columns[5]]="UnitPrice_F"
    return df.rename(columns=col_map)
def pipeline_3A_0_rename(df_bom,df_part_code,extras=None):
    if df_bom is None or df_bom.empty: return pd.DataFrame()
    df=df_bom.copy()
    if df_part_code is not None and not df_part_code.empty:
        rename_map=dict(zip(df_part_code.iloc[:,0].astype(str).str.strip(),df_part_code.iloc[:,1].astype(str).str.strip()))
        if "Original Type" in df.columns: df["Original Type"]=df["Original Type"].astype(str).str.strip().replace(rename_map)
    if "Original Article" not in df.columns: df["Original Article"]=df.iloc[:,0].astype(str)
    if extras: df=add_extra_components(df,[e for e in extras if e.get("target")=="bom"])
    return df
def pipeline_3A_1_filter(df_bom,df_stock):
    if df_bom is None or df_bom.empty: return pd.DataFrame()
    if df_stock is None or df_stock.empty: return df_bom.copy()
    cols=list(df_stock.columns)
    if len(cols)<3: return df_bom.copy()
    df_stock=df_stock.rename(columns={cols[0]:"Component",cols[2]:"Comment"}); excluded=df_stock[df_stock["Comment"].astype(str).str.lower().str.strip()=="no need"]["Component"].astype(str).str.upper().str.replace(" ","").str.strip().unique()
    df=df_bom.copy(); df["Norm_Type"]=df["Original Type"].astype(str).str.upper().str.replace(" ","").str.strip(); return df[~df["Norm_Type"].isin(excluded)].drop(columns=["Norm_Type"]).reset_index(drop=True)
def pipeline_3A_2_accessories(df_bom,df_acc):
    if df_acc is None or df_acc.empty: return df_bom
    out=df_bom.copy()
    for _,row in df_bom.iterrows():
        main_item=str(row["Original Type"]).strip(); matches=df_acc[df_acc.iloc[:,0].astype(str).str.strip()==main_item]
        for _,acc_row in matches.iterrows():
            acc_vals=acc_row.values[1:]
            for i in range(0,len(acc_vals),3):
                if i+2>=len(acc_vals) or pd.isna(acc_vals[i]): break
                item=str(acc_vals[i]).strip(); acc_qty=safe_parse_qty(str(acc_vals[i+1]).strip()); manuf=str(acc_vals[i+2]).strip()
                out=pd.concat([out,pd.DataFrame([{"Original Type":item,"Quantity":acc_qty,"Manufacturer":manuf,"Source":"Accessory"}])],ignore_index=True)
    return out
def pipeline_3A_3_nav(df_bom,df_part_no):
    if df_bom is None or df_bom.empty: return pd.DataFrame()
    if df_part_no is None or df_part_no.empty: df=df_bom.copy(); df["No."]=""; return df
    part=df_part_no.copy().reset_index(drop=True).rename(columns=lambda c:str(c).strip())
    if "PartName_B" not in part.columns or "PartNo_A" not in part.columns: df=df_bom.copy(); df["No."]=""; return df
    part["Norm_B"]=part["PartName_B"].astype(str).str.upper().str.replace(" ","").str.strip()
    def _n(x):
        try: return str(int(float(str(x).strip().replace(",","."))))
        except Exception: return str(x).strip()
    part["PartNo_A"]=part["PartNo_A"].map(_n).fillna("").astype(str); part=part.drop_duplicates(subset=["Norm_B"],keep="first").drop_duplicates(subset=["PartNo_A"],keep="first")
    df=df_bom.copy(); df["Norm_Type"]=df["Original Type"].astype(str).str.upper().str.replace(" ","").str.strip(); df["No."]=df["Norm_Type"].map(dict(zip(part["Norm_B"],part["PartNo_A"]))).fillna("").astype(str)
    merge_cols=[c for c in ["PartNo_A","Desc_C","Manufacturer_D","SupplierNo_E","UnitPrice_F","Norm_B"] if c in part.columns]
    if merge_cols:
        df=df.merge(part[merge_cols],left_on="No.",right_on="PartNo_A",how="left",suffixes=("","_dup")); df=df.rename(columns={"Desc_C":"Description","Manufacturer_D":"Supplier","SupplierNo_E":"Supplier No.","UnitPrice_F":"Unit Cost"})
        df=coalesce_cols(df,"Description",["Description_dup","Desc_C"]); df=coalesce_cols(df,"Supplier",["Supplier_dup"]); df=coalesce_cols(df,"Supplier No.",["Supplier No._dup"]); df=coalesce_cols(df,"Unit Cost",["Unit Cost_dup"]); df=df.drop(columns=[c for c in ["Norm_Type","Norm_B","PartNo_A"] if c in df.columns],errors="ignore")
    else: df=df.drop(columns=["Norm_Type"],errors="ignore")
    return ensure_scalar_strings(df)
def _read_stock_df(ks_file):
    if isinstance(ks_file,pd.DataFrame): stock=ks_file.copy()
    else: stock=pd.read_excel(io.BytesIO(ks_file.getvalue()),engine="openpyxl")
    stock=stock.rename(columns=lambda c:str(c).strip())
    cand_no=[c for c in stock.columns if c.lower() in ["no.","no","item no.","item no"]]; cand_bin=[c for c in stock.columns if c.lower() in ["bin code","bin","bin_code"]]; cand_qty=[c for c in stock.columns if c.lower() in ["quantity","qty","q"]]
    if cand_no and cand_bin and cand_qty: cols=[cand_no[0],cand_bin[0],cand_qty[0]]; stock=stock[cols]; stock.columns=["No.","Bin Code","Quantity"]
    else:
        cols=list(stock.columns)
        if len(cols)>=4: stock=stock[[cols[2],cols[1],cols[3]]]; stock.columns=["No.","Bin Code","Quantity"]
        else: return pd.DataFrame(columns=["No.","Bin Code","Quantity"])
    stock["No."]=stock["No."].apply(pipeline_1_4_normalize_no); stock["Quantity"]=pd.to_numeric(stock["Quantity"],errors="coerce").fillna(0.0); stock["Bin Code"]=stock["Bin Code"].astype(str).str.strip(); return stock
def pipeline_3A_4_stock(df_bom,ks_file):
    if df_bom is None or df_bom.empty: return pd.DataFrame()
    stock=_read_stock_df(ks_file); df=df_bom.copy(); df["No."]=df["No."].apply(pipeline_1_4_normalize_no); groups={k:v for k,v in stock.groupby("No.")}; df["Stock Rows"]=df["No."].map(groups); return df
def pipeline_3A_5_tables(df_bom,project_number,df_part_no):
    rows=[]
    for _,row in df_bom.iterrows():
        no=row.get("No."); qty=safe_parse_qty(row.get("Quantity",0)); stock_rows=row.get("Stock Rows")
        if not isinstance(stock_rows,pd.DataFrame) or stock_rows.empty: rows.append({"Entry Type":"Item","No.":no,"Document No.":f"{project_number}/N","Job No.":project_number,"Job Task No.":1144,"Quantity":qty,"Location Code":PURCHASE_LOCATION_CODE,"Bin Code":"","Description":row.get("Description",""),"Original Type":row.get("Original Type","")}); continue
        for alloc in allocate_from_stock(no,qty,stock_rows): rows.append({"Entry Type":"Item","No.":no,"Document No.":project_number,"Job No.":project_number,"Job Task No.":1144,"Quantity":alloc["Allocated Qty"],"Location Code":ALLOC_LOCATION_CODE if alloc["Bin Code"] else PURCHASE_LOCATION_CODE,"Bin Code":alloc["Bin Code"],"Description":row.get("Description",""),"Original Type":row.get("Original Type","")})
    job_journal=pd.DataFrame(rows); supplier_map=manuf_map={}
    if df_part_no is not None and not df_part_no.empty:
        if {"PartNo_A","SupplierNo_E"}.issubset(df_part_no.columns): supplier_map=dict(zip(df_part_no["PartNo_A"].astype(str),df_part_no["SupplierNo_E"]))
        if {"PartNo_A","Manufacturer_D"}.issubset(df_part_no.columns): manuf_map=dict(zip(df_part_no["PartNo_A"].astype(str),df_part_no["Manufacturer_D"].astype(str)))
    tmp=df_bom.copy()
    if "Quantity" not in tmp: tmp["Quantity"]=0
    if "Description" not in tmp: tmp["Description"]=""
    tmp["No."]=tmp["No."].astype(str); tmp["Quantity"]=pd.to_numeric(tmp["Quantity"],errors="coerce").fillna(0); nav_rows=[]
    for _,r in tmp.iterrows():
        part_no=str(r["No."]); qty=float(r.get("Quantity",0) or 0); manuf=(manuf_map or {}).get(part_no,""); profit=10 if "DANFOSS" in str(manuf).upper() else 17; supplier=(supplier_map or {}).get(part_no,30093); nav_rows.append({"Entry Type":"Item","No.":part_no,"Quantity":qty,"Supplier":supplier,"Profit":profit,"Discount":0,"Description":r.get("Description","")})
    nav_table=pd.DataFrame(nav_rows,columns=["Entry Type","No.","Quantity","Supplier","Profit","Discount","Description"]); return job_journal,nav_table,df_bom
def pipeline_3B_0_prepare_cubic(df_cubic,df_part_code,extras=None):
    if df_cubic is None or df_cubic.empty: return pd.DataFrame()
    df=df_cubic.copy().rename(columns=lambda c:str(c).strip())
    qty_cols=[c for c in df.columns if str(c).strip() in {"E","F","G"}]; combo=[c for c in df.columns if re.sub(r"\\s+","",str(c)).upper() in {"E+F+G","E+F","F+G","E+G"} or (("E" in str(c).upper()) and ("F" in str(c).upper()) and ("G" in str(c).upper()))]
    if qty_cols: df["Quantity"]=df[qty_cols].bfill(axis=1).iloc[:,0]
    elif combo:
        cc=combo[0]; df["Quantity"]=df[cc].apply(lambda v: safe_parse_qty(re.search(r"([0-9]+[.,]?[0-9]*)",str(v)).group(1)) if (pd.notna(v) and re.search(r"([0-9]+[.,]?[0-9]*)",str(v))) else 0.0)
    else:
        if "Quantity" not in df.columns: df["Quantity"]=0
    df["Quantity"]=pd.to_numeric(df["Quantity"],errors="coerce").fillna(0)
    if "Item Id" in df.columns: df["Original Type"]=df["Item Id"].astype(str).str.strip()
    elif "Original Type" not in df.columns: df["Original Type"]=df[df.columns[0]].astype(str)
    if "No." not in df.columns: df["No."]=df["Original Type"]
    if df_part_code is not None and not df_part_code.empty:
        rename_map=dict(zip(df_part_code.iloc[:,0].astype(str).str.strip(),df_part_code.iloc[:,1].astype(str).str.strip())); df["Original Type"]=df["Original Type"].astype(str).str.strip().replace(rename_map)
    if extras: df=add_extra_components(df,[e for e in extras if e.get("target")=="cubic"]); return df
def pipeline_3B_1_filtering(df_cubic,df_stock):
    if df_cubic is None or df_cubic.empty: return pd.DataFrame(),pd.DataFrame()
    if df_stock is None or df_stock.empty: return df_cubic.copy(),df_cubic.copy()
    cols=list(df_stock.columns)
    if len(cols)<3: return df_cubic.copy(),df_cubic.copy()
    s=df_stock.rename(columns={cols[0]:"Component",cols[2]:"Comment"}); bad=s[s["Comment"].astype(str).str.strip().str.lower().isin(["no need","q1"])]["Component"].astype(str); banned=bad.str.upper().str.replace(" ","").str.strip().unique()
    df=df_cubic.copy(); df["Norm_Type"]=df["Original Type"].astype(str).str.upper().str.replace(" ","").str.strip(); df_keep=df[~df["Norm_Type"].isin(banned)].reset_index(drop=True).drop(columns=["Norm_Type"]); return df_keep.copy(),df_keep.copy()
def pipeline_3B_2_accessories(df,df_acc):
    if df_acc is None or df_acc.empty: return df
    out=df.copy()
    for _,row in df.iterrows():
        main=str(row["Original Type"]).strip(); m=df_acc[df_acc.iloc[:,0].astype(str).str.strip()==main]; v=None
        for _,acc_row in m.iterrows():
            v=acc_row.values[1:]
            for i in range(0,len(v),3):
                if i+2>=len(v) or pd.isna(v[i]): break
                item=str(v[i]).strip(); qty=safe_parse_qty(str(v[i+1]).strip()); manuf=str(v[i+2]).strip(); out=pd.concat([out,pd.DataFrame([{"Original Type":item,"Quantity":qty,"Manufacturer":manuf,"Source":"Accessory"}])],ignore_index=True)
    return out
def pipeline_3B_3_nav(df,df_part_no): return pipeline_3A_3_nav(df,df_part_no)
def pipeline_3B_4_stock(df_journal,ks_file): return pipeline_3A_4_stock(df_journal,ks_file)
def pipeline_3B_5_tables(df_journal,df_nav,project_number,df_part_no):
    rows=[]
    for _,row in df_journal.iterrows():
        no=row.get("No."); qty=safe_parse_qty(row.get("Quantity",0)); stock_rows=row.get("Stock Rows")
        if not isinstance(stock_rows,pd.DataFrame) or stock_rows.empty: rows.append({"Entry Type":"Item","No.":no,"Document No.":f"{project_number}/N","Job No.":project_number,"Job Task No.":1144,"Quantity":qty,"Location Code":PURCHASE_LOCATION_CODE,"Bin Code":"","Description":row.get("Description",""),"Original Type":row.get("Original Type","")}); continue
        for alloc in allocate_from_stock(no,qty,stock_rows): rows.append({"Entry Type":"Item","No.":no,"Document No.":project_number,"Job No.":project_number,"Job Task No.":1144,"Quantity":alloc["Allocated Qty"],"Location Code":ALLOC_LOCATION_CODE if alloc["Bin Code"] else PURCHASE_LOCATION_CODE,"Bin Code":alloc["Bin Code"],"Description":row.get("Description",""),"Original Type":row.get("Original Type","")})
    job_journal=pd.DataFrame(rows); _,nav_table,_=pipeline_3A_5_tables(df_nav,project_number,df_part_no); return job_journal,nav_table,df_nav
def pipeline_4_1_calculation(df_bom,df_cubic,df_hours,panel_type,grounding,project_number,df_instr=None):
    if df_bom is None: df_bom=pd.DataFrame()
    if df_cubic is None: df_cubic=pd.DataFrame()
    if df_hours is None: df_hours=pd.DataFrame()
    if not df_bom.empty and {"Quantity","Unit Cost"}.issubset(df_bom.columns): parts_cost=(pd.to_numeric(df_bom["Quantity"],errors="coerce").fillna(0)*pd.to_numeric(df_bom["Unit Cost"],errors="coerce").fillna(0)).sum()
    else: parts_cost=0
    if not df_cubic.empty and {"Quantity","Unit Cost"}.issubset(df_cubic.columns): cubic_cost=(pd.to_numeric(df_cubic["Quantity"],errors="coerce").fillna(0)*pd.to_numeric(df_cubic["Unit Cost"],errors="coerce").fillna(0)).sum()
    else: cubic_cost=0
    hours_cost=0
    if not df_hours.empty and df_hours.shape[1]>4:
        hourly_rate=pd.to_numeric(df_hours.iloc[1,4],errors="coerce"); row=df_hours[df_hours.iloc[:,0].astype(str).str.upper()==str(panel_type).upper()]
        if not row.empty:
            if grounding=="TT": h=pd.to_numeric(row.iloc[0,1],errors="coerce")
            elif grounding=="TN-S": h=pd.to_numeric(row.iloc[0,2],errors="coerce")
            else: h=pd.to_numeric(row.iloc[0,3],errors="coerce")
            hours_cost=(h if pd.notna(h) else 0)*(hourly_rate if pd.notna(hourly_rate) else 0)
    smart_supply=9750.0; wire_set=2500.0; total=parts_cost+cubic_cost+hours_cost+smart_supply+wire_set; project_size=""; pallet_size=""
    if df_instr is not None and not df_instr.empty:
        row=df_instr[df_instr.iloc[:,0].astype(str).str.upper()==str(panel_type).upper()]
        if not row.empty: project_size=str(row.iloc[0,1]) if row.shape[1]>1 else ""; pallet_size=str(row.iloc[0,2]) if row.shape[1]>2 else ""
    return pd.DataFrame([{"Label":"Parts","Value":parts_cost},{"Label":"Cubic","Value":cubic_cost},{"Label":"Hours cost","Value":hours_cost},{"Label":"Smart supply","Value":smart_supply},{"Label":"Wire set","Value":wire_set},{"Label":"Extra","Value":0},{"Label":"Total","Value":total},{"Label":"Total+5%","Value":total*1.05},{"Label":"Total+35%","Value":total*1.35},{"Label":"Project size","Value":project_size},{"Label":"Pallet size","Value":pallet_size}])
def pipeline_4_2_missing_nav(df,source):
    if df is None or df.empty or "No." not in df.columns: return pd.DataFrame()
    missing=df[df["No."].astype(str).str.strip()=="" ] if not df.empty else pd.DataFrame()
    if missing.empty: return pd.DataFrame()
    qty=pd.to_numeric(missing.get("Quantity",0),errors="coerce").fillna(0).astype(float) if "Quantity" in missing else 0
    return pd.DataFrame({"Source":source,"Original Article":missing.get("Original Article",""),"Original Type":missing.get("Original Type",""),"Quantity":qty,"NAV No.":missing["No."]})

# ADD THIS HELPER (place it above render())
def run_processing(files, inputs):
    data_book = files.get("data", {})
    df_stock   = pipeline_2_3_get_sheet_safe(data_book, ["Stock"])
    df_part_no = pipeline_2_4_normalize_part_no(pipeline_2_3_get_sheet_safe(data_book, ["Part_no","Parts_no","Part no"]))
    df_hours   = pipeline_2_3_get_sheet_safe(data_book, ["Hours"])
    df_acc     = pipeline_2_3_get_sheet_safe(data_book, ["Accessories"])
    df_code    = pipeline_2_3_get_sheet_safe(data_book, ["Part_code"])
    df_instr   = pipeline_2_3_get_sheet_safe(data_book, ["Instructions"])
    extras=[]
    if inputs["ups"]:
        extras.extend([{"type":"LI32111CT01","qty":1,"target":"bom","force_no":"2214036"},
                       {"type":"ADV UPS holder V3","qty":1,"target":"bom","force_no":"2214035"},
                       {"type":"268-2610","qty":1,"target":"bom","force_no":"1865206"}])
    if inputs["swing_frame"]:
        extras.append({"type":"9030+2970","qty":1,"target":"cubic","force_no":"2185835"})
    if df_instr is not None and not df_instr.empty:
        row = df_instr[df_instr.iloc[:,0].astype(str).str.upper()==str(inputs["panel_type"]).upper()]
        if not row.empty:
            if inputs["panel_type"][0] not in ["F","G"]:
                try: qty_sdd = int(pd.to_numeric(row.iloc[0,4], errors="coerce").fillna(0))
                except Exception: qty_sdd = 0
                if qty_sdd>0: extras.append({"type":"SDD07550","qty":qty_sdd,"target":"cubic","force_no":"SDD07550"})
            for col_idx in range(5,10):
                if col_idx < row.shape[1]:
                    val = str(row.iloc[0,col_idx]).strip()
                    if val and val.lower()!="nan": extras.append({"type":val,"qty":1,"target":"cubic"})
    # Project BOM
    job_A = nav_A = df_bom_proc = pd.DataFrame()
    if all(k in files for k in ["bom","data","ks"]):
        df_bom = pipeline_3A_0_rename(files["bom"], df_code, extras)
        df_bom = pipeline_3A_1_filter(df_bom, df_stock)
        df_bom = pipeline_3A_2_accessories(df_bom, df_acc)
        df_bom = pipeline_3A_3_nav(df_bom, df_part_no)
        df_bom = pipeline_3A_4_stock(df_bom, files["ks"])
        job_A, nav_A, df_bom_proc = pipeline_3A_5_tables(df_bom, inputs["project_number"], df_part_no)
    # CUBIC BOM
    job_B = nav_B = df_cub_proc = pd.DataFrame()
    if (not inputs["rittal"]) and all(k in files for k in ["cubic_bom","data","ks"]):
        df_cubic = pipeline_3B_0_prepare_cubic(files["cubic_bom"], df_code, extras)
        df_j, df_n = pipeline_3B_1_filtering(df_cubic, df_stock)
        df_j = pipeline_3B_2_accessories(df_j, df_acc); df_n = pipeline_3B_2_accessories(df_n, df_acc)
        df_j = pipeline_3B_3_nav(df_j, df_part_no); df_n = pipeline_3B_3_nav(df_n, df_part_no)
        df_j = pipeline_3B_4_stock(df_j, files["ks"])
        job_B, nav_B, df_cub_proc = pipeline_3B_5_tables(df_j, df_n, inputs["project_number"], df_part_no)
    # Save once; later +/– clicks reuse this and avoid recompute
    st.session_state["proc"] = {"data_book":data_book,"df_stock":df_stock,"df_part_no":df_part_no,"df_hours":df_hours,"df_acc":df_acc,"df_code":df_code,"df_instr":df_instr,
                                "extras":extras,"job_A":job_A,"nav_A":nav_A,"df_bom_proc":df_bom_proc,"job_B":job_B,"nav_B":nav_B,"df_cub_proc":df_cub_proc}

# IN render(), REPLACE the block from the Run Processing button down to the end of "3B" processing with this:
    if st.button("🚀 Run Processing"):
        st.session_state["processing_started"] = True
        st.session_state["mech_confirmed"] = False
        st.session_state["df_mech"] = pd.DataFrame()
        st.session_state["df_remain"] = pd.DataFrame()
        run_processing(files, inputs)
    if not st.session_state.get("processing_started", False): st.stop()
    # Reuse cached processing; prevents app “reset” on +/– clicks
    if "proc" not in st.session_state: run_processing(files, inputs)
    proc = st.session_state["proc"]
    df_stock = proc["df_stock"]; df_part_no = proc["df_part_no"]; df_hours = proc["df_hours"]; df_acc = proc["df_acc"]; df_code = proc["df_code"]; df_instr = proc["df_instr"]
    job_A = proc["job_A"]; nav_A = proc["nav_A"]; df_bom_proc = proc["df_bom_proc"]
    job_B = proc["job_B"]; nav_B = proc["nav_B"]; df_cub_proc = proc["df_cub_proc"]
    # inject CSS only once to avoid duplicate headers/logos feeling
    if not st.session_state.get("css_injected"):
        st.markdown("<style>.app {font-family:system-ui}</style>", unsafe_allow_html=True)
        st.session_state["css_injected"] = True


def render():
    st.header(f"BOM Management · {get_app_version()}")
    inputs=pipeline_2_1_user_inputs()
    if not inputs: return
    st.session_state["inputs"]=inputs
    files=pipeline_2_2_file_uploads(inputs["rittal"]); 
    # keep uploads cached; do not reset on +/- reruns
    if "processing_started" not in st.session_state: st.session_state["processing_started"] = False
    if "mech_confirmed" not in st.session_state: st.session_state["mech_confirmed"] = False
    if "df_mech" not in st.session_state: st.session_state["df_mech"] = pd.DataFrame()
    if "df_remain" not in st.session_state: st.session_state["df_remain"] = pd.DataFrame()
    # launch processing once; do not clear cached uploads on rerun
    if st.button("🚀 Run Processing"):
        st.session_state["processing_started"] = True
        st.session_state["mech_confirmed"] = False
        st.session_state["df_mech"] = pd.DataFrame()
        st.session_state["df_remain"] = pd.DataFrame()
        st.session_state.pop("export_bundle", None)
    if not st.session_state["processing_started"]:
        st.stop()
    if not files: return
    reqA=["bom","data","ks"]; reqB=["cubic_bom","data","ks"] if not inputs["rittal"] else []; missA=[k for k in reqA if k not in files]; missB=[k for k in reqB if k not in files]
    st.subheader("📋 Required files"); c1,c2=st.columns(2)
    with c1: st.success("Project BOM: OK") if not missA else st.warning(f"Project BOM missing: {missA}")
    with c2: st.success("CUBIC BOM: OK") if (not inputs["rittal"] and not missB) else (st.warning(f"CUBIC BOM missing: {missB}") if not inputs["rittal"] else st.info("CUBIC BOM skipped (Rittal)"))
    if st.button("🚀 Run Processing"): st.session_state.update({"processing_started":True,"mech_confirmed":False,"df_mech":pd.DataFrame(),"df_remain":pd.DataFrame()}); st.session_state.pop("export_bundle",None)
    if not st.session_state.get("processing_started",False): st.stop()
    data_book=files.get("data",{}); df_stock=pipeline_2_3_get_sheet_safe(data_book,["Stock"]); df_part_no=pipeline_2_4_normalize_part_no(pipeline_2_3_get_sheet_safe(data_book,["Part_no","Parts_no","Part no"])); df_hours=pipeline_2_3_get_sheet_safe(data_book,["Hours"]); df_acc=pipeline_2_3_get_sheet_safe(data_book,["Accessories"]); df_code=pipeline_2_3_get_sheet_safe(data_book,["Part_code"]); df_instr=pipeline_2_3_get_sheet_safe(data_book,["Instructions"])
    extras=[]
    if inputs["ups"]: extras.extend([{"type":"LI32111CT01","qty":1,"target":"bom","force_no":"2214036"},{"type":"ADV UPS holder V3","qty":1,"target":"bom","force_no":"2214035"},{"type":"268-2610","qty":1,"target":"bom","force_no":"1865206"}])
    if inputs["swing_frame"]: extras.append({"type":"9030+2970","qty":1,"target":"cubic","force_no":"2185835"})
    if df_instr is not None and not df_instr.empty:
        row=df_instr[df_instr.iloc[:,0].astype(str).str.upper()==str(inputs["panel_type"]).upper()]
        if not row.empty:
            if inputs["panel_type"][0] not in ["F","G"]:
                try: qty_sdd=int(pd.to_numeric(row.iloc[0,4],errors="coerce").fillna(0))
                except Exception: qty_sdd=0
                if qty_sdd>0: extras.append({"type":"SDD07550","qty":qty_sdd,"target":"cubic","force_no":"SDD07550"})
            for cidx in range(5,10):
                if cidx<row.shape[1]:
                    v=str(row.iloc[0,cidx]).strip()
                    if v and v.lower()!="nan": extras.append({"type":v,"qty":1,"target":"cubic"})
    _norm_type=lambda s: str(s).upper().replace(" ","").strip()
    def _norm_no(x):
        sx=str(x).strip()
        try: return str(int(float(sx.replace(",","."))))
        except: return sx
    def _get_ex(df_stock_):
        if df_stock_ is None or df_stock_.empty or df_stock_.shape[1]<3: return set(),set()
        cols=list(df_stock_.columns); s=df_stock_.rename(columns={cols[0]:"Component",cols[2]:"Comment"}); m=s["Comment"].astype(str).str.strip().str.lower().isin(["no need","q1"]); comp=s.loc[m,"Component"].astype(str); return set(comp.map(_norm_type)),set(comp.map(_norm_no))
    ex_type,ex_no=_get_ex(df_stock)
    def _apply_excl(df):
        if df is None or df.empty: return df
        t=df.copy()
        if "Original Type" in t.columns:
            t["_T"]=t["Original Type"].map(_norm_type); t["_N"]=t["No."].map(_norm_no) if "No." in t.columns else ""
            t=t[~t["_T"].isin(ex_type) & (~t["_N"].isin(ex_no) if "No." in t.columns else True)].drop(columns=["_T","_N"],errors="ignore")
        return t
    job_A=nav_A=df_bom_proc=pd.DataFrame()
    if not missA:
        df_bom=pipeline_3A_0_rename(files["bom"],df_code,extras); df_bom=pipeline_3A_1_filter(df_bom,df_stock); df_bom=pipeline_3A_2_accessories(df_bom,df_acc); df_bom=pipeline_3A_3_nav(df_bom,df_part_no); df_bom=pipeline_3A_4_stock(df_bom,files["ks"]); job_A,nav_A,df_bom_proc=pipeline_3A_5_tables(df_bom,inputs["project_number"],df_part_no)
    job_B=nav_B=df_cub_proc=pd.DataFrame()
    if not inputs["rittal"] and not missB:
        df_cubic=pipeline_3B_0_prepare_cubic(files["cubic_bom"],df_code,extras); df_j,df_n=pipeline_3B_1_filtering(df_cubic,df_stock); df_j=pipeline_3B_2_accessories(df_j,df_acc); df_n=pipeline_3B_2_accessories(df_n,df_acc); df_j=pipeline_3B_3_nav(df_j,df_part_no); df_n=pipeline_3B_3_nav(df_n,df_part_no); df_j=pipeline_3B_4_stock(df_j,files["ks"]); job_B,nav_B,df_cub_proc=pipeline_3B_5_tables(df_j,df_n,inputs["project_number"],df_part_no)
    if not st.session_state.get("mech_confirmed",False):
        if not job_B.empty:
            st.subheader("📑 Job Journal (CUBIC BOM → allocate to Mechanics)")
            st.markdown("""
            <style>
            .mech-table{margin-top:4px}
            .mech-table *{font-family:system-ui,Segoe UI,Arial,sans-serif!important;color:#fff!important;font-weight:500!important}
            .mech-row{display:flex;align-items:center}
            .mech-cell{display:flex;align-items:center;min-height:40px;padding:4px 8px}
            .mech-hr{border:0;border-top:1px solid rgba(255,255,255,.25);margin:2px 0 0}
            .qty-box{display:flex;align-items:center;justify-content:center;gap:8px}
            .qty-display{min-width:60px;text-align:center;font-weight:800!important;font-size:18px;padding:2px 10px;border:1px solid rgba(255,255,255,.35);border-radius:8px}
            .btn-darkgreen .stButton>button{background:#074b22!important;color:#fff!important;border:none!important;height:38px!important;width:38px!important;min-width:38px!important;border-radius:8px!important;padding:0!important;line-height:36px!important;font-size:18px!important;font-weight:800!important;box-shadow:none!important}
            .btn-placeholder{display:inline-block;height:38px;width:38px;border-radius:8px;border:1px solid rgba(255,255,255,.22)}
            </style>
            """, unsafe_allow_html=True)
            st.session_state.setdefault("mech_take",{})
            editable=_apply_excl(job_B.copy()); editable["Available Qty"]=editable["Quantity"].astype(float)
            if editable.empty:
                st.info("No selectable items (filtered by Stock comments: No need/Q1)."); st.session_state["mech_confirmed"]=True; st.stop()
            # header
            h=st.columns([3,3,6,1.0,3])
            h[0].markdown("**No.**"); h[1].markdown("**Original Type**"); h[2].markdown("**Description**"); h[3].markdown("**Qty**"); h[4].markdown("**Allocate**")
            st.markdown("<div class='mech-table'>", unsafe_allow_html=True)
            def _inc(k,mx): st.session_state["mech_take"][k]=min(st.session_state["mech_take"].get(k,0.0)+1,mx)
            def _dec(k): st.session_state["mech_take"][k]=max(st.session_state["mech_take"].get(k,0.0)-1,0.0)
            for idx,row in editable.iterrows():
                cols=st.columns([3,3,6,1.0,3])
                with cols[0]: st.markdown(f"<div class='mech-cell'>{str(row.get('No.',''))}</div>", unsafe_allow_html=True)
                with cols[1]: st.markdown(f"<div class='mech-cell'>{str(row.get('Original Type',''))}</div>", unsafe_allow_html=True)
                with cols[2]: st.markdown(f"<div class='mech-cell'>{str(row.get('Description',''))}</div>", unsafe_allow_html=True)
                with cols[3]: st.markdown(f"<div class='mech-cell'>{int(row.get('Quantity',0))}</div>", unsafe_allow_html=True)
                with cols[4]:
                    key=f"take_{idx}"; mx=float(row["Available Qty"]); cur=float(st.session_state["mech_take"].get(key,0.0))
                    b=st.columns([1,2,1])
                    with b[0]:
                        if cur>0:
                            with st.container(): st.markdown("<div class='btn-darkgreen'>", unsafe_allow_html=True); st.button("−", key=f"minus_{idx}", on_click=_dec, args=(key,), use_container_width=True); st.markdown("</div>", unsafe_allow_html=True)
                        else: st.markdown("<span class='btn-placeholder'></span>", unsafe_allow_html=True)
                    with b[1]: st.markdown(f"<div class='mech-cell qty-box'><div class='qty-display'>{cur:.0f}</div></div>", unsafe_allow_html=True)
                    with b[2]:
                        if cur<mx:
                            with st.container(): st.markdown("<div class='btn-darkgreen'>", unsafe_allow_html=True); st.button("+", key=f"plus_{idx}", on_click=_inc, args=(key,mx), use_container_width=True); st.markdown("</div>", unsafe_allow_html=True)
                        else: st.markdown("<span class='btn-placeholder'></span>", unsafe_allow_html=True)
                st.markdown("<hr class='mech-hr'/>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            # (confirm button logic below remains unchanged)
            st.session_state.setdefault("mech_take",{})
            editable=_apply_excl(job_B.copy()); editable["Available Qty"]=editable["Quantity"].astype(float)
            if editable.empty:
                st.info("No selectable items (filtered by Stock comments: No need/Q1)."); st.session_state["mech_confirmed"]=True; st.stop()
            # header
            h=st.columns([3,3,6,1.2,3])
            h[0].markdown("**No.**"); h[1].markdown("**Original Type**"); h[2].markdown("**Description**"); h[3].markdown("**Qty**"); h[4].markdown("**Allocate**")
            st.markdown("<div class='mech-zone'>", unsafe_allow_html=True)
            def _inc(k,mx): st.session_state["mech_take"][k]=min(st.session_state["mech_take"].get(k,0.0)+1,mx)
            def _dec(k): st.session_state["mech_take"][k]=max(st.session_state["mech_take"].get(k,0.0)-1,0.0)
            for idx,row in editable.iterrows():
                cols=st.columns([3,3,6,1.2,3])
                with cols[0]: st.markdown(f"<div class='mech-cell'><p class='mech-label'>{str(row.get('No.',''))}</p></div>", unsafe_allow_html=True)
                with cols[1]: st.markdown(f"<div class='mech-cell'><p class='mech-label'>{str(row.get('Original Type',''))}</p></div>", unsafe_allow_html=True)
                with cols[2]: st.markdown(f"<div class='mech-cell'><p class='mech-label'>{str(row.get('Description',''))}</p></div>", unsafe_allow_html=True)
                with cols[3]: st.markdown(f"<div class='mech-cell'><p class='mech-label'>{int(row.get('Quantity',0))}</p></div>", unsafe_allow_html=True)
                with cols[4]:
                    key=f"take_{idx}"; mx=float(row["Available Qty"]); cur=float(st.session_state["mech_take"].get(key,0.0))
                    bcols=st.columns([1,2,1])
                    with bcols[0]:
                        if cur>0: st.button("−", key=f"minus_{idx}", on_click=_dec, args=(key,), use_container_width=True)
                        else: st.markdown("<span class='btn-placeholder'></span>", unsafe_allow_html=True)
                    with bcols[1]: st.markdown(f"<div class='mech-cell qty-box'><div class='qty-display'>{cur:.0f}</div></div>", unsafe_allow_html=True)
                    with bcols[2]:
                        if cur<mx: st.button("+", key=f"plus_{idx}", on_click=_inc, args=(key,mx), use_container_width=True)
                        else: st.markdown("<span class='btn-placeholder'></span>", unsafe_allow_html=True)
                st.markdown("<hr class='mech-hr'/>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            if st.button("✅ Confirm Mechanics Allocation"):
                mech_rows,remain_rows=[],[]
                for idx,row in editable.iterrows():
                    k=f"take_{idx}"; take=float(st.session_state["mech_take"].get(k,0.0)); avail=float(row["Available Qty"]); r=row.to_dict()
                    if take>0: mech_rows.append({**r,"Quantity":take})
                    rem=max(avail-take,0.0)
                    if rem>0 and str(r.get("No.",""))!="2185835": remain_rows.append({**r,"Quantity":rem})
                st.session_state["df_mech"]=pd.DataFrame(mech_rows); st.session_state["df_remain"]=pd.DataFrame(remain_rows); st.session_state["mech_confirmed"]=True
                if inputs["swing_frame"]:
                    swing=pd.DataFrame([{"Entry Type":"Item","Original Type":"9030+2970","No.":"2185835","Quantity":1,"Document No.":inputs["project_number"],"Job No.":inputs["project_number"],"Job Task No.":1144,"Location Code":PURCHASE_LOCATION_CODE,"Bin Code":"","Description":"Swing frame component","Source":"Extra"}]); st.session_state["df_mech"]=pd.concat([st.session_state["df_mech"],swing],ignore_index=True)
        st.stop()
    def _show(df,title): 
        if df is not None and not df.empty: st.subheader(title); st.data_editor(df,use_container_width=True,hide_index=True,height=300)
    st.session_state["df_mech"]=_apply_excl(st.session_state.get("df_mech")); st.session_state["df_remain"]=_apply_excl(st.session_state.get("df_remain"))
    _show(st.session_state.get("df_mech"),"📑 Job Journal (CUBIC BOM TO MECH.)"); _show(st.session_state.get("df_remain"),"📑 Job Journal (CUBIC BOM REMAINING)"); _show(job_A,"📑 Job Journal (Project BOM)"); _show(nav_A,"🛒 NAV Table (Project BOM)"); _show(nav_B,"🛒 NAV Table (CUBIC BOM)")
    calc=pipeline_4_1_calculation(df_bom_proc,df_cub_proc,df_hours,inputs["panel_type"],inputs["grounding"],inputs["project_number"],df_instr); _show(calc,"💰 Calculation")
    miss_nav_A=pipeline_4_2_missing_nav(df_bom_proc,"Project BOM"); miss_nav_B=pipeline_4_2_missing_nav(df_cub_proc,"CUBIC BOM"); _show(miss_nav_A,"⚠️ Missing NAV Numbers (Project BOM)"); _show(miss_nav_B,"⚠️ Missing NAV Numbers (CUBIC BOM)")
    st.session_state["export_bundle"]={"inputs":inputs,"calc":calc,"job_A":job_A,"nav_A":nav_A,"job_B":job_B,"nav_B":nav_B,"miss_nav_A":miss_nav_A,"miss_nav_B":miss_nav_B,"df_mech":st.session_state.get("df_mech"),"df_remain":st.session_state.get("df_remain")}
    st.subheader("💾 Export")
    if st.button("💾 Export Results to Excel"):
        b=st.session_state.get("export_bundle",{})
        if not b: st.warning("Nothing to export – run processing first."); st.stop()
        ts=datetime.datetime.now().strftime("%Y%m%d%H%M")
        try: project_size=str(b["calc"][b["calc"]["Label"]=="Project size"]["Value"].iloc[0]); pallet_size=str(b["calc"][b["calc"]["Label"]=="Pallet size"]["Value"].iloc[0])
        except Exception: project_size=pallet_size=""
        filename=f"{b['inputs']['project_number']}_{b['inputs']['panel_type']}_{b['inputs']['grounding']}_{pallet_size}_{ts}.xlsx"
        wb=Workbook(); ws=wb.active; ws.title="Info"; info=[["Project number",b["inputs"]["project_number"]],["Panel type",b["inputs"]["panel_type"]],["Grounding",b["inputs"]["grounding"]],["Main switch",b["inputs"]["main_switch"]],["Swing frame",b["inputs"]["swing_frame"]],["UPS",b["inputs"]["ups"]],["Rittal",b["inputs"]["rittal"]],["Project size",project_size],["Pallet size",pallet_size]]
        for r in info: ws.append(r)
        ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=20
        bold=Font(bold=True); grey=PatternFill(start_color="DDDDDD",end_color="DDDDDD",fill_type="solid"); thin=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
        for r in ws["A1":"A9"]:
            for c in r: c.font=bold; c.fill=grey; c.border=thin
        for r in ws["B1":"B9"]:
            for c in r: c.border=thin
        def add_df_to_wb(df,title,colw=None,nav=False,calc=False):
            if df is None or df.empty: return
            df=ensure_scalar_strings(df); w=wb.create_sheet(title); w.append(df.columns.tolist())
            for _,row in df.iterrows(): w.append(list(row.values))
            if colw:
                for col,wid in colw.items(): w.column_dimensions[col].width=wid
            mr,mc=w.max_row,w.max_column
            for rr in w.iter_rows(min_row=1,max_row=mr,min_col=1,max_col=mc):
                for cc in rr: cc.border=thin
            if nav:
                for rr in w["A1":"G1"]:
                    for cc in rr: cc.font=bold; c.fill=grey
            if calc:
                for rr in w["A1":"A10"]:
                    for cc in rr: cc.font=bold; cc.fill=grey
                for rr in w["B2":"B10"]:
                    for cc in rr: cc.number_format=CURRENCY_FORMAT
        job_w={"A":8,"B":10,"C":12,"D":12,"E":12,"F":12,"G":13,"H":12,"I":40,"J":25}
        add_df_to_wb(b["df_mech"],"JobJournal_Mech",job_w); add_df_to_wb(b["df_remain"],"JobJournal_Remaining",job_w); add_df_to_wb(b["job_A"],"JobJournal_ProjectBOM",job_w); add_df_to_wb(b["job_B"],"JobJournal_CUBICBOM",job_w)
        nav_w={"A":8,"B":10,"C":9,"D":9,"E":9,"F":9,"G":50}
        add_df_to_wb(b["nav_B"],"NAV_CUBICBOM",nav_w,nav=True); add_df_to_wb(b["nav_A"],"NAV_ProjectBOM",nav_w,nav=True); add_df_to_wb(b["calc"],"Calculation",{"A":12,"B":18},calc=True); add_df_to_wb(b["miss_nav_A"],"MissingNAV_ProjectBOM"); add_df_to_wb(b["miss_nav_B"],"MissingNAV_CUBICBOM")
        path=f"/mnt/data/{filename}"; wb.save(path); st.download_button("⬇️ Download Excel",data=open(path,"rb"),file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
if __name__=="__main__": render()