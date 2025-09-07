import streamlit as st
import pandas as pd
import io, re
from collections import Counter
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="ADV Management", layout="wide")

def norm_name(x):
    return ''.join(str(x).upper().split())

def parse_qty(x):
    if pd.isna(x): return 0.0
    if isinstance(x,(int,float)): return float(x)
    s=str(x).strip().replace('\xa0','').replace(' ','')
    if ',' in s and '.' in s: s=s.replace(',','')
    else: s=s.replace('.','').replace(',','.')
    try: return float(s)
    except: return 0.0

def safe_filename(s):
    s='' if s is None else str(s).strip()
    s=re.sub(r'[\\/:*?"<>|]+','',s)
    return s.replace(' ','_')

def build_kaunas_index(df_kaunas):
    idx={}
    for _, r in df_kaunas.sort_values(['Component','Bin Code']).iterrows():
        key=r['Norm']; q=float(r['Quantity'])
        if q<=0: continue
        idx.setdefault(key,[]).append([r['Bin Code'], q])
    return idx

def allocate_bins_for_table(df_orders, bins_index):
    if df_orders is None or df_orders.empty:
        return pd.DataFrame(columns=list(df_orders.columns)+['Bin Code'])
    rows=[]
    for _, r in df_orders.iterrows():
        item=r['Item No.']; key=norm_name(item); remaining=float(r['Quantity'])
        if remaining<=0: continue
        if key not in bins_index:
            row=r.copy(); row['Document No.']=str(row.get('Document No.',''))+'/NERA'; row['Bin Code']=''; rows.append(row); continue
        bins=bins_index[key]; i=0
        while remaining>0 and i<len(bins):
            bin_code, avail=bins[i]
            if avail<=0: i+=1; continue
            take=min(remaining, avail)
            row=r.copy(); row['Job Task No.']=row.get('Job Task No.',1144); row['Quantity']=round(take,2); row['Bin Code']=bin_code
            rows.append(row); remaining-=take; bins[i][1]=round(avail-take,6)
            if bins[i][1]<=1e-9: i+=1
        if remaining>1e-9:
            row=r.copy(); row['Quantity']=round(remaining,2); row['Document No.']=str(row.get('Document No.',''))+'/NERA'; row['Bin Code']=''; rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=list(df_orders.columns)+['Bin Code'])

def process_stock_usage_keep_zeros(df_target, stock_df):
    used_rows, adjusted_rows = [], []
    stock_tmp=stock_df.copy()
    stock_tmp['Norm']=stock_tmp['Component'].astype(str).map(norm_name)
    stock_norm_qty=stock_tmp.groupby('Norm',as_index=True)['Quantity'].sum().to_dict()
    for _, row in df_target.iterrows():
        item=row['Item No.']; key=norm_name(item); qty_needed=float(row['Quantity']); qty_used=0.0
        qty_in_stock=float(stock_norm_qty.get(key,0))
        if qty_in_stock>0:
            qty_used=min(qty_needed, qty_in_stock); stock_norm_qty[key]=qty_in_stock-qty_used
        if qty_used>0: used_rows.append({'Item No.': item, 'Used from stock': qty_used})
        qty_remaining=qty_needed-qty_used
        adjusted=row.copy(); adjusted['Job Task No.']=adjusted.get('Job Task No.',1144); adjusted['Quantity']=round(qty_remaining,2)
        adjusted_rows.append(adjusted)
    return pd.DataFrame(adjusted_rows), pd.DataFrame(used_rows)

def move_no_second_item_last(df):
    if df is None or df.empty: return df
    cols=list(df.columns)
    if 'No.' in cols: cols.insert(1, cols.pop(cols.index('No.')))
    if 'Item No.' in cols: cols.append(cols.pop(cols.index('Item No.')))
    return df[cols]

def finalize_bom_alloc_table(df, vendor_to_no=None):
    if df is None or df.empty: return df
    out=df.copy()
    if 'Used from stock' in out.columns: out=out.drop(columns=['Used from stock'])
    if 'Item No.' in out.columns: out=out.rename(columns={'Item No.': 'Vendor Item Number'})
    if 'Cross-Reference No.' in out.columns: out=out.rename(columns={'Cross-Reference No.': 'Vendor Item Number'})
    if 'No.' not in out.columns and 'Vendor Item Number' in out.columns:
        if vendor_to_no is not None: out['No.']=out['Vendor Item Number'].map(lambda v: vendor_to_no.get(norm_name(v)))
        else: out['No.']=out['Vendor Item Number'].map(lambda v: _part_b2a.get(norm_name(v)))
    order=['Type','No.','Document No.','Job No.','Job Task No.','Quantity','Location Code','Bin Code','Vendor Item Number']
    out=out[[c for c in order if c in out.columns]+[c for c in out.columns if c not in order]]
    return move_no_second_item_last(out)

def reorder_supplier_after_discount(df):
    if df is None or df.empty: return df
    cols=list(df.columns); changed=False
    if 'Supplier' in cols and 'Discount' in cols: cols.remove('Supplier'); cols.insert(cols.index('Discount')+1,'Supplier'); changed=True
    if 'Supplier No.' in cols and 'Discount' in cols: cols.remove('Supplier No.'); cols.insert(cols.index('Discount')+1,'Supplier No.'); changed=True
    return df[cols] if changed else df

def get_main_switch_and_accessories(ms_df, selected):
    if ms_df is None or ms_df.empty or not selected: return []
    row=ms_df[ms_df.iloc[:,0].astype(str).str.strip().str.upper()==str(selected).strip().upper()]
    if row.empty: return []
    vals=[str(selected).strip()]+[str(v).strip() for v in row.iloc[0,1:].tolist() if pd.notna(v) and str(v).strip()]
    seen=set(); out=[]
    for v in vals:
        key=norm_name(v)
        if key and key not in seen: seen.add(key); out.append(v)
    return out

st.title("ADV Management Tool")

file_cubic=st.file_uploader("Įkelkite CUBIC.xls", type=["xls"])
file_bom=st.file_uploader("Įkelkite BOM list", type=["xlsx","xls"])
file_data=st.file_uploader("Įkelkite Data.xlsx", type=["xlsx"])
file_ks=st.file_uploader("Įkelkite Kaunas Stock Excel", type=["xlsx","xls"])

doc_number=st.text_input("Document No.")
type_choice=st.selectbox("Tipas",['A','B','B1','B2','C','C1','C2','C3','C4','C4.1','C5','C6','C7','C8','F','F1','F2','F3','F4','F4.1','F5','F6','F7','G','G1','G2','G3','G4','G5','G6','G7','Custom'])
main_switch_choice=st.selectbox("Kirtiklis",["C160S4FM","C125S4FM","C080S4FM","31115","31113","31111","31109","31107","C404400S","C634630S"])
earthing_choice=st.selectbox("Įžeminimas",['TT','TN-S','TN-C-S'])
swing=st.checkbox("Ar bus Swing Frame?")
ups=st.checkbox("Ar bus UPS?")

generate=st.button("✅ Generuoti")

if generate and file_cubic and file_bom and file_data and file_ks and doc_number:
    try:
        df_raw=pd.read_excel(file_cubic, engine="xlrd")
        df_cubic=df_raw.iloc[14:,[1,2,4]]
        df_cubic.columns=['Vendor','Description','Quantity']
        df_cubic=df_cubic.dropna(subset=['Vendor']).copy()
        df_cubic['Type']='item'; df_cubic['Supplier']=30093; df_cubic['Profit']=17; df_cubic['Discount']=0
        df_cubic=df_cubic[['Type','Vendor','Quantity','Supplier','Profit','Discount','Description']]
        df_stock_comments=pd.read_excel(file_data, sheet_name="Stock", usecols=[0,1,2], names=["Component","Quantity","Comment"], header=None, skiprows=1)
        df_stock_comments['Component']=df_stock_comments['Component'].astype(str)
        df_stock_available=df_stock_comments.copy()
        df_stock_available['Quantity']=pd.to_numeric(df_stock_available['Quantity'], errors='coerce').fillna(0)
        df_stock_available=df_stock_available[df_stock_available['Quantity']>0]
        df_instructions=pd.read_excel(file_data, sheet_name="Instructions", usecols=[0,2,3,4,5,6,7,8,9], header=0, names=["Type","Palette","Image","Quantity","Comp1","Comp2","Comp3","Comp4","Comp5"])
        df_part_no=pd.read_excel(file_data, sheet_name='Part_no', usecols="A:F", header=0)
        df_part_no.columns=['PartNo_A','PartName_B','Desc_C','Manufacturer_D','SupplierNo_E','UnitPrice_F']
        df_part_no['Norm_B']=df_part_no['PartName_B'].astype(str).map(norm_name)
        global _part_b2a; _part_b2a=dict(zip(df_part_no['Norm_B'], df_part_no['PartNo_A']))
        df_kaunas=pd.read_excel(file_ks, sheet_name=0, header=None, skiprows=3, usecols=[1,3,10])
        df_kaunas.columns=['Bin Code','Quantity','Component']
        df_kaunas=df_kaunas.dropna(subset=['Component']).copy()
        df_kaunas['Quantity']=df_kaunas['Quantity'].apply(parse_qty).fillna(0)
        df_kaunas['Norm']=df_kaunas['Component'].apply(norm_name)
        df_kaunas=df_kaunas[df_kaunas['Quantity']>0].reset_index(drop=True)
        kaunas_bins_index=build_kaunas_index(df_kaunas)
        st.success("Failai įkelti, duomenys apdoroti")
        excluded_vendors=df_stock_comments[df_stock_comments["Comment"].isin(["Q1","No need","Wurth","GRM"])]["Component"].dropna().unique()
        df_filtered=df_cubic[~df_cubic["Vendor"].isin(excluded_vendors)].reset_index(drop=True)
        df_filtered=df_filtered.iloc[1:].reset_index(drop=True)
        df_filtered['Quantity']=pd.to_numeric(df_filtered['Quantity'], errors='coerce').fillna(0)

        st.subheader("Mechanikos kiekių įvedimas")
        mech_entries=[]
        for i,row in df_filtered.iterrows():
            qty=st.number_input(f"{row['Vendor']} ({row['Description']}) [Max {int(row['Quantity'])}]",0,int(row['Quantity']),0,key=f"mech_{i}")
            if qty>0:
                mech_entries.append({
                    'Type':row['Type'],
                    'Document No.':doc_number,
                    'Job No.':doc_number,
                    'Job Task No.':1144,
                    'Item No.':row['Vendor'],
                    'Quantity':float(qty),
                    'Location Code':'KAUNAS'
                })

        df_mech=pd.DataFrame(mech_entries)
        if swing:
            df_mech=pd.concat([df_mech,pd.DataFrame([{
                'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,
                'Item No.':'9030+2970','Quantity':1,'Location Code':'KAUNAS'
            }])],ignore_index=True)

        if not df_mech.empty:
            df_mech,df_used_mech=process_stock_usage_keep_zeros(df_mech,df_stock_available)
            df_mech=df_mech[df_mech['Quantity']>0].copy()
            mech_alloc=allocate_bins_for_table(df_mech,kaunas_bins_index)
            if not mech_alloc.empty:
                mech_alloc['No.']=mech_alloc['Item No.'].map(lambda v:_part_b2a.get(norm_name(v),v))
                mech_alloc=move_no_second_item_last(mech_alloc)
        else:
            mech_alloc=pd.DataFrame(); df_used_mech=pd.DataFrame()

        df_remaining=df_filtered.copy()
        df_remaining['Kiekis įvestas']=[st.session_state[f"mech_{i}"] for i in range(len(df_filtered))]
        df_remaining['Likęs kiekis']=df_remaining['Quantity']-df_remaining['Kiekis įvestas']
        df_base=pd.DataFrame({
            'Type':df_remaining['Type'],
            'Document No.':doc_number,
            'Job No.':doc_number,
            'Job Task No.':1144,
            'Item No.':df_remaining['Vendor'],
            'Quantity':df_remaining['Likęs kiekis'],
            'Location Code':'KAUNAS'
        })

        df_expanded=[]
        for _,r in df_base.iterrows():
            rr=r.copy()
            if 'PLC' in str(rr['Item No.']).upper() and type_choice.startswith(('A','B','C')):
                rr['Item No.']='9030+0040'
            rr['Quantity']=round(float(rr['Quantity']),2)
            df_expanded.append(rr)
        df_result=pd.DataFrame(df_expanded)

        if ups:
            df_result=pd.concat([df_result,pd.DataFrame([{
                'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,
                'Item No.':'ADV UPS HOLDER V3','Quantity':1,'Location Code':'KAUNAS'
            }])],ignore_index=True)
        if swing:
            df_result=pd.concat([df_result,pd.DataFrame([
                {'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,'Item No.':'1055-1000','Quantity':2,'Location Code':'KAUNAS'},
                {'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,'Item No.':'1055-1001','Quantity':2,'Location Code':'KAUNAS'}
            ])],ignore_index=True)

        df_result=pd.concat([df_result,pd.DataFrame([
            {'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,'Item No.':'THERMOSTAT_PLATE_HOL','Quantity':2,'Location Code':'KAUNAS'},
            {'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,'Item No.':'DOOR_BACK_BRACKET_4','Quantity':5,'Location Code':'KAUNAS'}
        ])],ignore_index=True)

        match=df_instructions[df_instructions['Type']==type_choice]
        if not match.empty and pd.notna(match.iloc[0]['Quantity']):
            qty_sdd=int(match.iloc[0]['Quantity'])
            if qty_sdd>0:
                df_result=pd.concat([df_result,pd.DataFrame([{
                    'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,
                    'Item No.':'SDD07550','Quantity':qty_sdd,'Location Code':'KAUNAS'
                }])],ignore_index=True)

        if not match.empty:
            comp_cols=['Comp1','Comp2','Comp3','Comp4','Comp5']
            comps=[str(match.iloc[0][c]).strip() for c in comp_cols if pd.notna(match.iloc[0][c]) and str(match.iloc[0][c]).strip()]
            comp_counts=Counter(comps)
            for item,qty in comp_counts.items():
                df_result=pd.concat([df_result,pd.DataFrame([{
                    'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,
                    'Item No.':item,'Quantity':qty,'Location Code':'KAUNAS'
                }])],ignore_index=True)

        df_result,df_used_prod=process_stock_usage_keep_zeros(df_result,df_stock_available)
        df_result=df_result[df_result['Quantity']>0].copy()
        prod_alloc=allocate_bins_for_table(df_result,kaunas_bins_index)
        if not prod_alloc.empty:
            prod_alloc['No.']=prod_alloc['Item No.'].map(lambda v:_part_b2a.get(norm_name(v),v))
            prod_alloc=move_no_second_item_last(prod_alloc)

        if df_used_prod.empty: df_used_prod=pd.DataFrame()
        if df_used_mech.empty: df_used_mech=pd.DataFrame()
        df_used_total=pd.concat([df_used_prod,df_used_mech],ignore_index=True).groupby('Item No.',as_index=False)['Used from stock'].sum()

        df_part_code=pd.read_excel(file_data,sheet_name='Part_code')
        df_accessories=pd.read_excel(file_data,sheet_name='Accessories')
        df_main_switch=pd.read_excel(file_data,sheet_name='Main_switch',usecols="B:H")

        df_stock_simple=pd.read_excel(file_data,sheet_name='Stock',usecols=[0,1])
        df_stock_simple.columns=['Item ref.','Stock Quantity']
        df_stock_simple['Item ref.']=df_stock_simple['Item ref.'].astype(str).str.strip()

        df_bom_raw=pd.read_excel(file_bom)
        df_bom=df_bom_raw.copy()
        df_bom.columns=['Original_Item_Ref','Type','Quantity','Manufacturer','Description']
        df_partcode_full=pd.read_excel(file_data,sheet_name='Part_code',usecols=[0,1,2],header=0)
        df_partcode_full.columns=['From','To','PC_Manufacturer']
        name_map={norm_name(a):str(b).strip() for a,b in df_partcode_full[['From','To']].dropna(subset=['From']).values if pd.notna(b) and str(b).strip()}
        manuf_map={}
        for _,r in df_partcode_full.iterrows():
            a,b,m=r['From'],r['To'],r['PC_Manufacturer']
            if pd.isna(a): continue
            if pd.notna(m) and str(m).strip():
                if pd.notna(b) and str(b).strip(): manuf_map[norm_name(str(b))]=str(m).strip()
                manuf_map.setdefault(norm_name(str(a)),str(m).strip())
        df_bom['Type']=df_bom['Type'].apply(lambda s:name_map.get(norm_name(s),s) if (pd.notna(s) and str(s).strip()) else s)
        mask_missing=df_bom['Manufacturer'].isna()|(df_bom['Manufacturer'].astype(str).str.strip()=='')
        df_bom.loc[mask_missing,'Manufacturer']=df_bom.loc[mask_missing,'Type'].apply(lambda s:manuf_map.get(norm_name(s)) if (pd.notna(s) and str(s).strip()) else None)

        selected_ms=main_switch_choice
        ms_items=get_main_switch_and_accessories(df_main_switch,selected_ms)
        if ms_items:
            df_ms=pd.DataFrame([{
                'Type':'item','Document No.':doc_number,'Job No.':doc_number,'Job Task No.':1144,
                'Item No.':it,'Quantity':1,'Location Code':'KAUNAS'
            } for it in ms_items])
            ms_alloc=allocate_bins_for_table(df_ms,kaunas_bins_index)
            if not ms_alloc.empty:
                ms_alloc=ms_alloc.rename(columns={'Item No.':'Vendor Item Number'})
                ms_alloc=finalize_bom_alloc_table(ms_alloc)
        else:
            ms_alloc=pd.DataFrame()

        exclude_manufacturers=["BITZER KÜHLMACHINENBAU GMBH","BELDEN","KABELTEC","HELUKABEL","LAPP","GENERAL CAVI","EMERSON CLIMATE TECHNOLOGIES","ELFAC A/S","DEKA CONTROLS GMBH","TYPE SPECIFIED IN BOM","PRYSMIAN","CO4","BELIMO","PEPPERL + FUCHS","WAGO"]
        exclude_types=["47KOHM","134F7613","134H7160","CABLE JZ","AKSF","ROUNDPACKART","XALK178E","LMBWLB32-180S","ACH580-01-026A-4","ACH580-01-033A-4","ACH580-01-073A-4","PSP650MT3-230U"]
        comments_to_exclude={"Q1","NO NEED"}
        _df_sc=df_stock_comments.copy()
        _df_sc["Component_norm"]=_df_sc["Component"].astype(str).str.strip().str.upper()
        _df_sc["Comment_norm"]=_df_sc["Comment"].astype(str).str.strip().str.upper()
        exclude_by_comment=set(_df_sc.loc[_df_sc["Comment_norm"].isin(comments_to_exclude),"Component_norm"].unique())
        _dfm=df_bom.copy()
        _dfm["Final_Item_Ref_norm"]=_dfm["Type"].astype(str).str.strip().str.upper()
        _dfm["Final_Manufacturer_norm"]=_dfm["Manufacturer"].astype(str).str.strip().str.upper()
        mask_types=~_dfm["Final_Item_Ref_norm"].isin(set(t.upper() for t in exclude_types))
        mask_comment=~_dfm["Final_Item_Ref_norm"].isin(exclude_by_comment)
        mask_manuf=~_dfm["Final_Manufacturer_norm"].isin(set(m.upper() for m in exclude_manufacturers))
        df_bom_filtered=_dfm[mask_types & mask_comment & mask_manuf].reset_index(drop=True)

        sms_alloc=pd.DataFrame()
        if not df_bom_filtered.empty:
            schneider=df_bom_filtered[(df_bom_filtered['Manufacturer']=="Schneider Electric")&(df_bom_filtered['Type'].astype(str).isin(pd.unique(df_main_switch.values.ravel())))]
            other=df_bom_filtered[(df_bom_filtered['Manufacturer']!="WAGO")&~((df_bom_filtered['Manufacturer']=="Schneider Electric")&(df_bom_filtered['Type'].astype(str).isin(pd.unique(df_main_switch.values.ravel()))))]
            if not schneider.empty:
                sms_src=schneider.rename(columns={'Type':'Item No.'}).copy()
                sms_src['Quantity']=pd.to_numeric(sms_src['Quantity'],errors='coerce').fillna(0).astype(float)
                sms_alloc=allocate_bins_for_table(sms_src,kaunas_bins_index)
                if not sms_alloc.empty:
                    sms_alloc=sms_alloc.rename(columns={'Item No.':'Vendor Item Number'})
                    sms_alloc=finalize_bom_alloc_table(sms_alloc,_part_b2a)
            if not other.empty:
                other_src=other.rename(columns={'Type':'Item No.'}).copy()
                other_src['Quantity']=pd.to_numeric(other_src['Quantity'],errors='coerce').fillna(0).astype(float)
                other_alloc=allocate_bins_for_table(other_src,kaunas_bins_index)
                if not other_alloc.empty:
                    other_alloc=other_alloc.rename(columns={'Item No.':'Vendor Item Number'})
                    other_alloc=finalize_bom_alloc_table(other_alloc,_part_b2a)
        else:
            other_alloc=pd.DataFrame()

        palette_val=''
        match_for_name=df_instructions[df_instructions['Type']==type_choice]
        if not match_for_name.empty and pd.notna(match_for_name.iloc[0]['Palette']):
            palette_val=str(match_for_name.iloc[0]['Palette']).strip()
        today_str=pd.Timestamp.today().strftime('%Y-%m-%d')
        export_filename=f"{safe_filename(doc_number)}_{safe_filename(type_choice)}_{safe_filename(earthing_choice)}_{safe_filename(palette_val)}_{today_str}.xlsx"

        with pd.ExcelWriter(export_filename,engine="openpyxl") as writer:
            if not mech_alloc.empty: mech_alloc.to_excel(writer,sheet_name="Mechanika",index=False)
            if not prod_alloc.empty: prod_alloc.to_excel(writer,sheet_name="Gamyba",index=False)
            if not df_used_total.empty: df_used_total.to_excel(writer,sheet_name="Paimta iš Stock",index=False)
            if not sms_alloc.empty: sms_alloc.to_excel(writer,sheet_name="BOM Schneider Main Switch",index=False)
            if not other_alloc.empty: other_alloc.to_excel(writer,sheet_name="BOM Likę Komponentai",index=False)

            df_hours=pd.read_excel(file_data,sheet_name="Hours",header=None)
            hourly_rate=float(df_hours.iloc[1,4])
            proj_type=type_choice.upper(); earthing=earthing_choice.upper()
            row_match=df_hours[df_hours.iloc[:,0].astype(str).str.upper()==proj_type]
            hours_value=0
            if not row_match.empty:
                if earthing=="TT": hours_value=float(row_match.iloc[0,1])
                elif earthing=="TN-S": hours_value=float(row_match.iloc[0,2])
                elif earthing=="TN-C-S": hours_value=float(row_match.iloc[0,3])
            hours_cost=hours_value*hourly_rate
            smart_supply_cost=9750.0; wire_set_cost=2500.0
            parts_cost=10000; cubic_cost=5000
            total_cost=parts_cost+cubic_cost+hours_cost+smart_supply_cost+wire_set_cost

            wb=writer.book
            ws=wb.create_sheet('Calculation')
            ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=20
            thin=Side(style='thin'); border_all=Border(left=thin,right=thin,top=thin,bottom=thin); bold=Font(bold=True)
            ws['A2']=f"{doc_number} | {proj_type} | {earthing} | {palette_val}"; ws['A2'].font=bold
            rows=[("Parts:",parts_cost,""),("Cubic:",cubic_cost,""),("Hours cost:",hours_cost,f"{int(hours_value)} Hours"),("Smart supply:",smart_supply_cost,""),("Wire set:",wire_set_cost,""),("Extra:",None,""),("Total:",None,""),("Total+5%:",None,""),("Total+35%:",None,"")]
            start_row=4
            for r,(label,value,info) in enumerate(rows,start=start_row):
                ws.cell(row=r,column=1,value=label)
                if label.startswith("Total"): ws.cell(row=r,column=1).font=bold
                if label=="Extra:": ws.cell(row=r,column=2,value=None)
                elif value is not None: ws.cell(row=r,column=2,value=float(value))
                else: ws.cell(row=r,column=2,value=None)
                ws.cell(row=r,column=3,value=info if info else None)
            B=lambda row:f"B{row}"; total_row=start_row+6
            ws[B(total_row)]=f"=SUM(B{start_row}:B{start_row+5})"
            ws[B(total_row+1)]=f"={B(total_row)}*1.05"
            ws[B(total_row+2)]=f"={B(total_row)}*1.35"
            for r in range(start_row,start_row+9):
                val=ws[B(r)].value
                ws[B(r)].number_format='#,##0.00 "DKK"'
            for rr in (total_row,total_row+1,total_row+2):
                ws[f"A{rr}"].font=bold; ws[f"B{rr}"].font=bold
                for c in range(1,4): ws.cell(row=rr,column=c).border=border_all; ws.cell(row=rr,column=c).alignment=Alignment(vertical='center')

        with open(export_filename,"rb") as f:
            st.download_button("📥 Atsisiųsti Excel rezultatą",f,file_name=export_filename)

    except Exception as e:
        st.error(f"❌ Klaida apdorojant failus: {e}")
